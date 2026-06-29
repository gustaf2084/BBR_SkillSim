# -*- coding: utf-8 -*-
"""
build_data_json.py
将《战场兄弟·重铸》0.7.6 的 xlsx 转换为 data.json（供 BBR_SkillSimulator 使用）。

数据来源工作表：
  - Backgrounds            背景 ID、骰组修正、权重倍率、专属组
  - Perk Group Probabilities  背景 × 技能树组 的基础出现概率矩阵 (100×47)
  - Tree Multipliers       特性→组权重 / 天赋星→组权重 / 专属组级联权重 / 投影属性
  - Perks                  各技能树组 Tier1–7 技能清单（流派推荐用）

输出：data.json，结构见开发计划文档第四节。

关键决策（已与用户确认）：
  A. Tactician 组无概率矩阵列 → 基础概率记为 null，出现概率由权重规则合成。
  B. 专属组 None=不骰组(级联为空)；+=强加成(权重 9999)；混合模式按概率分支 + 强加成分开。
  C. 组名以概率矩阵 47 组为规范键；大小写统一；战斗风格用矩阵名(Power/Ranged/Swift/Shield)。

用法：
  python build_data_json.py <xlsx路径> <输出data.json路径>
  python build_data_json.py   # 默认读取同目录 bb.xlsx，输出 data.json
"""

import json
import os
import re
import sys
from collections import OrderedDict

import openpyxl

# ---------------------------------------------------------------------------
# 常量与规范
# ---------------------------------------------------------------------------

BIG_WEIGHT = 10.0           # 特性/专属组 '+' 强加成的近似权重（调低以保留概率区分度，避免饱和到100%）
EPS = 1e-9

# 概率矩阵的 47 个组名（规范键，col2..col48）
CANON_GROUPS_47 = [
    "Agile", "Fast", "Tough", "Trained", "Unstoppable", "Vicious", "Vigorous",
    "Axe", "Bow", "Cleaver", "Crossbow", "Dagger", "Flail", "Hammer", "Mace",
    "Polearm", "Spear", "Sword", "Throwing",
    "Heavy Armor", "Light Armor", "Medium Armor",
    "Swift", "Power", "Ranged", "Shield",
    "Knave", "Laborer", "Militia", "Noble", "Pauper", "Raider", "Soldier",
    "Swordmaster", "Trapper", "Wildling",
    "Back To Basics", "Discovered Talent", "Fencer", "Gifted", "Leadership",
    "Man Of Steel", "Marksmanship", "Professional", "Rising Star", "Student",
    "General",
]

# 组 → 类别映射（按动态骰组顺序）
GROUP_CATEGORY = {}
for g in ["Agile", "Fast", "Tough", "Trained", "Unstoppable", "Vicious", "Vigorous", "Tactician"]:
    GROUP_CATEGORY[g] = "Shared"
for g in ["Knave", "Laborer", "Militia", "Noble", "Pauper", "Raider", "Soldier",
          "Swordmaster", "Trapper", "Wildling"]:
    GROUP_CATEGORY[g] = "Exclusive"
for g in ["Axe", "Bow", "Cleaver", "Crossbow", "Dagger", "Flail", "Hammer", "Mace",
          "Polearm", "Spear", "Sword", "Throwing"]:
    GROUP_CATEGORY[g] = "Weapon"
for g in ["Heavy Armor", "Light Armor", "Medium Armor"]:
    GROUP_CATEGORY[g] = "Armor"
for g in ["Swift", "Power", "Ranged", "Shield"]:
    GROUP_CATEGORY[g] = "Fighting Style"
for g in ["Back To Basics", "Discovered Talent", "Fencer", "Gifted", "Leadership",
          "Man Of Steel", "Marksmanship", "Professional", "Rising Star", "Student"]:
    GROUP_CATEGORY[g] = "Special"
GROUP_CATEGORY["General"] = "Always"

# 组名归一化映射：将 Perks 表 / Tree Multipliers 中的变体名 → 规范键
GROUP_NAME_NORMALIZE = {
    "Back to Basics": "Back To Basics",
    "Man of Steel": "Man Of Steel",
    "Powerful Strikes": "Power",
    "Ranged Combat": "Ranged",
    "Swift Strikes": "Swift",
    "Swift\nStrikes": "Swift",
    "Leadership (special)": "Leadership",
    "Leadership\n(special)": "Leadership",
}

# 骰组修正关键词 → 类别
ROLL_KEYWORD_TO_CATEGORY = {
    "Shared": "Shared",
    "Exclusive": "Exclusive",
    "Weapon": "Weapon",
    "Armor": "Armor",
    "Fighting Style": "Fighting Style",
    "Special": "Special",
    "Special Perks": "Special",
}

# 默认骰组数量基线（来自 Backgrounds 表 col2 表头）
DEFAULT_GROUP_ROLLS = {
    "Exclusive": (0, 1),   # 0-1：范围，取期望 0.5
    "Shared": 2,
    "Weapon": 3,
    "Armor": 2,
    "Fighting Style": 2,
    "Special": 0,          # Special 默认骰组数未在表头给出，按 0
}

# 远程类组（Melee Only 背景禁用）
RANGED_GROUPS = {"Bow", "Crossbow", "Throwing", "Ranged"}

# 护甲类别简写 → 规范组名（col7 中 "Medium +/Light +/Heavy +" 的简写）
ARMOR_SHORTCUT = {
    "Light": "Light Armor",
    "Medium": "Medium Armor",
    "Heavy": "Heavy Armor",
}

# 数据中的非标准组名 → 规范组名（笔误/变体）
GROUP_ALIAS = {
    "Agility": "Agile",
    "Talented": "Trained",
    "Devious": "Knave",
}

# 自由文本说明（应跳过，非组名/权重）
FREE_TEXT_TOKENS = {
    "remaning two", "remaining two", "50 each", "40%", "60%", "remove professional",
    "each", "two", "remaning", "remaining",
}

# ---------------------------------------------------------------------------
# 解析告警收集
# ---------------------------------------------------------------------------
WARNINGS = []


def warn(msg):
    WARNINGS.append(msg)
    print(f"  [WARN] {msg}")


def info(msg):
    print(f"  [info] {msg}")


# ---------------------------------------------------------------------------
# 通用解析工具
# ---------------------------------------------------------------------------

def norm_group(name):
    """归一化组名为规范键。未知组名返回 None 并告警（自由文本也返回 None）。"""
    if name is None:
        return None
    s = str(name).strip()
    if not s:
        return None
    # 剥离括号说明（如 "Swordmaster (all perks are on t4...)" → "Swordmaster"）
    s = re.sub(r"\s*\([^)]*\)\s*", "", s).strip()
    if not s:
        return None
    # 自由文本说明 → 跳过
    if s.lower() in FREE_TEXT_TOKENS:
        return None
    if s in GROUP_NAME_NORMALIZE:
        return GROUP_NAME_NORMALIZE[s]
    if s in GROUP_ALIAS:
        return GROUP_ALIAS[s]
    if s in GROUP_CATEGORY:
        return s
    # 大小写容错
    for k in GROUP_CATEGORY:
        if s.lower() == k.lower():
            return k
    # 护甲简写
    if s in ARMOR_SHORTCUT:
        return ARMOR_SHORTCUT[s]
    if s == "None":
        return "None"  # 专属组占位，特殊处理
    if s == "Melee Only":
        return None    # 标记，不是组
    # 类别标题（Shared/Armor/Weapon 等纯标题行）静默跳过，不告警
    if s in ("Shared", "Armor", "Weapon", "Fighting Style", "Special Perks",
             "Special Groups", "Group", "Exclusive", "Projected Attributes"):
        return None
    warn(f"未知组名(忽略): {repr(s)}")
    return None


def parse_weight_value(token):
    """解析单个权重值: '+'→BIG_WEIGHT, 数字→float, ''→None。"""
    t = token.strip()
    if t == "" :
        return None
    if t == "+":
        return BIG_WEIGHT
    if t.lower() == "0":
        return 0.0
    try:
        return float(t)
    except ValueError:
        return None  # 静默返回 None（配对失败时由调用方处理）


def parse_weight_pairs(tokens):
    """
    将已拆分的 token 列表按 '组, 权重' 两两配对。
    tokens: ['Agile','4','Fast','4', ...] 或含 '组 +' / 单独组名。
    返回 OrderedDict {group: weight}，并收集强加成组与无法配对项。
    """
    result = OrderedDict()
    forced = []
    i = 0
    n = len(tokens)
    while i < n:
        tok = tokens[i].strip()
        if not tok:
            i += 1
            continue
        gname = norm_group(tok)
        # 下一项是权重值
        if i + 1 < n:
            nxt = tokens[i + 1].strip()
            if nxt == "+":
                if gname:
                    result[gname] = BIG_WEIGHT
                i += 2
                continue
            wval = parse_weight_value(nxt)
            if wval is not None:
                if gname and gname != "None":
                    result[gname] = wval
                # gname 为 None(自由文本) 或 "None" 时，连数值一起跳过
                i += 2
                continue
        # 当前 token 形如 "Agile 4"（同一格内空格分隔）
        m = re.match(r"^([A-Za-z][A-Za-z ]*?)\s+([+\d.]+)$", tok)
        if m:
            gname2 = norm_group(m.group(1))
            w2 = parse_weight_value(m.group(2))
            if gname2 and gname2 != "None" and w2 is not None:
                result[gname2] = w2
            i += 1
            continue
        # 单独组名（无权重）→ 强加成
        if gname and gname != "None":
            result[gname] = BIG_WEIGHT
        i += 1
    return result


def parse_weight_list(text):
    """
    解析多行/逗号分隔的 '组, 权重' 列表（用于特性、multipliers 等单列文本）。
    自动剔除 'Melee Only' / 'Weight Groups' 等非组标记。
    """
    if not text:
        return OrderedDict()
    s = str(text).strip()
    if not s:
        return OrderedDict()
    # 剔除整行标记（这些标记独占一行或作为段落标题）
    for marker in ["Weight Groups", "Melee Only"]:
        s = s.replace(marker, " ")
    # 把 "组 +" 无逗号写法规范为 "组, +"（如 "Medium +" → "Medium, +"）
    s = re.sub(r"\b([A-Za-z]+)\s+\+", r"\1, +", s)
    # "组 数值" 无逗号写法规范为 "组, 数值"（如 "Vicious 1.5" → "Vicious, 1.5"）
    s = re.sub(r"\b([A-Za-z]+)\s+(\d[\d.]*)", r"\1, \2", s)
    s = s.replace("\n", ",").replace("/", ",")
    s = re.sub(r"\band\b", ",", s, flags=re.IGNORECASE)
    tokens = [t.strip() for t in s.split(",") if t.strip()]
    return parse_weight_pairs(tokens)


def to_float(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (ValueError, AttributeError):
        return default


# ---------------------------------------------------------------------------
# 各表解析
# ---------------------------------------------------------------------------

def parse_prob_matrix(ws):
    """Perk Group Probabilities 表 → {background_id: {group: base_prob}}"""
    hdr = list(ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column, values_only=True))[0]
    col_groups = []
    for h in hdr[1:]:
        if h and str(h).strip():
            col_groups.append(norm_group(h))
        else:
            col_groups.append(None)
    info(f"概率矩阵: {sum(1 for g in col_groups if g)} 个有效组列")

    matrix = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue
        bg = str(row[0]).strip()
        probs = {}
        for idx, val in enumerate(row[1:], start=0):
            g = col_groups[idx]
            if g is None:
                continue
            p = to_float(val, None)
            if p is not None:
                probs[g] = p
        matrix[bg] = probs
    info(f"概率矩阵: {len(matrix)} 个背景")
    return matrix, col_groups


def clean_bg_id(raw_id):
    """背景 ID 去除括号后缀，如 'paladin_background (oathtaker)' → 'paladin_background'"""
    s = str(raw_id).strip()
    # 去掉括号及之后内容
    s = re.sub(r"\s*\([^)]*\)\s*", "", s).strip()
    return s


def parse_group_rolls(roll_raw, melee_only_flag):
    """
    解析 col2 'Groups Rolled' 修正，返回 {category: 增减量}。
    仅记录背景对默认基线的 +/- 修正；默认基线在引擎层叠加。
    """
    group_rolls = {}
    if not roll_raw:
        return group_rolls
    s = str(roll_raw).strip()
    # 形如 "+1 Shared", "-1 Fighting Style", "+2 Weapon", 多个用空格/换行分隔
    for m in re.finditer(r"([+-]\d+)\s*([A-Za-z ]+)", s):
        num = int(m.group(1))
        cat_raw = m.group(2).strip()
        # 大小写不敏感匹配类别
        cat = None
        for kw, c in ROLL_KEYWORD_TO_CATEGORY.items():
            if cat_raw.lower() == kw.lower():
                cat = c
                break
        if cat:
            group_rolls[cat] = group_rolls.get(cat, 0) + num
        else:
            warn(f"未识别骰组类别 '{cat_raw}' (raw={s})")
    return group_rolls


def parse_backgrounds(ws, matrix):
    """Backgrounds 表 → {bg_id: {...}}。权重数据合并 col3..col9 多列。"""
    bgs = OrderedDict()
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue
        bg_raw = str(row[0]).strip()
        bg = clean_bg_id(bg_raw)
        roll_raw = str(row[1]).strip() if row[1] else ""              # col2 Groups Rolled
        # col3..col9: multipliers / Exclusive / Shared / Weapon / Armor / Fighting Style / Special Perks
        # col4 是 Exclusive（单独处理）；col3,5,6,7,8,9 是各类别权重
        col_mult   = row[2]   # col3
        col_shared = row[4]   # col5
        col_weapon = row[5]   # col6
        col_armor  = row[6]   # col7
        col_style  = row[7]   # col8
        col_special = row[8]  # col9
        exc_raw = str(row[3]).strip() if row[3] else ""               # col4 Exclusive

        # Melee Only 检测：可能在 col3 或 col6
        melee_only = ("Melee Only" in (str(col_mult) if col_mult else "")) or \
                     ("Melee Only" in (str(col_weapon) if col_weapon else ""))

        # 合并所有权重列
        multipliers = OrderedDict()
        for col in (col_mult, col_shared, col_weapon, col_armor, col_style, col_special):
            if col and str(col).strip():
                parsed = parse_weight_list(col)
                for g, w in parsed.items():
                    multipliers[g] = w  # 后者覆盖，但通常不冲突

        # 骰组修正
        group_rolls = parse_group_rolls(roll_raw, melee_only)

        # 专属组
        exclusive = parse_exclusive(exc_raw, bg)

        # 基础概率（按清洗后 ID 从矩阵取）
        base_probs = matrix.get(bg, {})
        if not base_probs and bg not in matrix:
            warn(f"背景 {bg} (raw={bg_raw}) 在概率矩阵中无对应行，基础概率为空")

        bgs[bg] = OrderedDict([
            ("display_zh", ""),
            ("group_rolls", group_rolls),
            ("multipliers", multipliers),
            ("melee_only", melee_only),
            ("exclusive", exclusive),
            ("base_probabilities", base_probs),
            ("icon", f"{bg}.png"),
        ])
    info(f"背景表: 解析 {len(bgs)} 个背景")
    return bgs


def parse_exclusive(exc_raw, bg):
    """
    解析专属组列 col4，返回结构:
      {"mode": "none"}                         无专属组
      {"mode": "fixed", "group": "Noble"}      固定专属组(可能含 +)
      {"mode": "prob", "picks": {组:概率}, "none_chance": float}
      {"mode": "mixed", "forced": [组], "picks": {组:概率}, "none_chance": float}
    """
    if not exc_raw:
        return {"mode": "none"}
    # 说明性文本(如 hedge_knight 的 "(Lone Wolf perk on tier 1)")
    if exc_raw.startswith("("):
        warn(f"背景 {bg}: 专属组为说明性文本，视为 none: {exc_raw[:40]}")
        return {"mode": "none"}

    s = exc_raw.replace("\n", ",").replace("/", ",")
    s = re.sub(r"\band\b", ",", s, flags=re.IGNORECASE)
    tokens = [t.strip() for t in s.split(",") if t.strip()]

    forced = []          # 强加成组(+)
    picks = {}           # 概率组: 组->百分比
    none_chance = 0.0
    i = 0
    while i < len(tokens):
        tok = tokens[i]
        # 形如 "组, 数值"
        if i + 1 < len(tokens) and re.match(r"^[A-Za-z ]+$", tok):
            grp = norm_group(tok)
            nxt = tokens[i + 1].strip()
            if nxt == "+":
                forced.append(grp)
                i += 2
                continue
            try:
                pct = float(nxt)
                if grp.lower() == "none":
                    none_chance += pct
                else:
                    picks[grp] = picks.get(grp, 0.0) + pct
                i += 2
                continue
            except ValueError:
                pass
        # 单独的组名(无数值) → 视为强加成
        if re.match(r"^[A-Za-z ]+$", tok):
            grp = norm_group(tok)
            forced.append(grp)
            i += 1
            continue
        i += 1

    has_prob = bool(picks or none_chance)
    has_forced = bool(forced)
    if has_forced and not has_prob:
        return {"mode": "fixed", "group": forced[0]}
    if has_prob and not has_forced:
        return {"mode": "prob", "picks": picks, "none_chance": none_chance}
    if has_prob and has_forced:
        return {"mode": "mixed", "forced": forced, "picks": picks, "none_chance": none_chance}
    return {"mode": "none"}

def parse_traits(ws):
    """Tree Multipliers 左侧特性 → {trait_name: {weights, display_zh, icon}}"""
    traits = OrderedDict()
    # 特性从 R13 开始，到 R 列 A 连续为空或遇到非特性区域为止。
    # 已知特性区: R13..R89 (77个)
    for i in range(13, 200):
        r = list(ws.iter_rows(min_row=i, max_row=i, min_col=1, max_col=3, values_only=True))[0]
        name = r[0]
        if not name or not str(name).strip():
            # 连续空行后停止（但允许中间偶发空行？实际特性区连续）
            if traits:
                # 再看下一行是否还有
                nxt = list(ws.iter_rows(min_row=i + 1, max_row=i + 1, min_col=1, max_col=1, values_only=True))[0][0]
                if not nxt:
                    break
                else:
                    continue
            else:
                continue
        sname = str(name).strip()
        # 跳过小节标题(如 "Character Traits")
        if sname in ("Character Traits", "Talents", "Characteristic"):
            continue
        weights = parse_weight_list(r[2])
        if not weights and sname not in ("Addict", "Arena Fighter", "Arena Pit Fighter",
                                          "Arena Veteran", "Cultist Acolyte", "Cultist Chosen",
                                          "Cultist Disciple", "Cultist Fanatic", "Cultist Prophet",
                                          "Cultist Zealot", "Fear Beasts", "Fear Greenskins",
                                          "Fear Nobles", "Fear Undead", "Glorious Endurance",
                                          "Glorious Quickness", "Glorious Resolve", "Hate Beasts",
                                          "Hate Greenskins", "Hate Nobles", "Hate Undead",
                                          "Intensive Training", "Irrational", "Loyal", "Lucky",
                                          "Mad", "Night Owl", "Old", "Player Character",
                                          "Teamplayer", "Deathwish", "Determined"):
            # 无权重且非已知"无权重"特性 → 仍保留(权重为空)
            pass
        traits[sname] = OrderedDict([
            ("display_zh", ""),
            ("weights", weights),
            ("icon", f"{sname}.png"),
        ])
    info(f"特性表: 解析 {len(traits)} 个特性")
    return traits


def parse_attribute_stars(ws):
    """Tree Multipliers 左侧天赋星(R3..R10) → {attribute: {group: weight}}"""
    attrs = OrderedDict()
    # R3 Hitpoints .. R10 Ranged Defense
    attr_rows = {
        3: "Hitpoints", 4: "Bravery", 5: "Fatigue", 6: "Initiative",
        7: "Melee Skill", 8: "Ranged Skill", 9: "Melee Defense", 10: "Ranged Defense",
    }
    for i, aname in attr_rows.items():
        r = list(ws.iter_rows(min_row=i, max_row=i, min_col=1, max_col=3, values_only=True))[0]
        w = parse_weight_list(r[2])
        if w:
            attrs[aname] = w
    return attrs


def parse_exclusive_cascade(ws):
    """Tree Multipliers 右侧专属组(R3..R12) → {group: {other_weights}}"""
    casc = OrderedDict()
    for i in range(3, 13):
        r = list(ws.iter_rows(min_row=i, max_row=i, min_col=5, max_col=7, values_only=True))[0]
        name = r[0]
        if not name or not str(name).strip():
            continue
        gname = norm_group(name)
        other_w = parse_weight_list(r[2])
        casc[gname] = {"other_weights": other_w}
    return casc


def parse_self_weights(ws):
    """Tree Multipliers 右侧各组 self_weight(R15..R58) → {group: self_weight}"""
    sw = {}
    for i in range(15, 62):
        r = list(ws.iter_rows(min_row=i, max_row=i, min_col=5, max_col=6, values_only=True))[0]
        name = r[0]
        if not name or not str(name).strip():
            continue
        gname = norm_group(name)
        if gname in ("Shared", "Armor", "Weapon", "Fighting Style", "Special Perks",
                     "Special Groups", "Projected Attributes", "Group", "Exclusive"):
            continue
        val = r[1]
        v = to_float(val, None)
        if v is not None:
            sw[gname] = v
    return sw


def parse_projected_attributes(ws):
    """Tree Multipliers 右侧投影属性(R73..R87) → {group: {base, cutoff}}"""
    proj = {}
    for i in range(73, 88):
        r = list(ws.iter_rows(min_row=i, max_row=i, min_col=5, max_col=7, values_only=True))[0]
        name = r[0]
        if not name or not str(name).strip():
            continue
        gname = norm_group(name)
        base = to_float(r[1], None)
        cutoff = to_float(r[2], None)
        if base is not None:
            proj[gname] = {"base": base, "cutoff": cutoff if cutoff is not None else None}
    return proj


def parse_perks(ws):
    """Perks 表 → {group: {Tier1..Tier7: skill_desc}}"""
    trees = OrderedDict()
    current = None
    for i in range(1, ws.max_row + 1):
        v = list(ws.iter_rows(min_row=i, max_row=i, min_col=1, max_col=1, values_only=True))[0][0]
        if not v or not str(v).strip():
            continue
        s = str(v).strip()
        if s.startswith("PERK GROUP CATEGORY"):
            current = None
            continue
        if s == "Perk Group" or s == "Perk Name":
            continue
        gname = norm_group(s)
        if gname in GROUP_CATEGORY:
            # 这是一个组行，读 Tier1..Tier7 (col2..col8)
            tiers = {}
            trow = list(ws.iter_rows(min_row=i, max_row=i, min_col=2, max_col=8, values_only=True))[0]
            for ti, tv in enumerate(trow, start=1):
                if tv and str(tv).strip():
                    tiers[f"Tier {ti}"] = str(tv).strip()
            trees[gname] = tiers
            current = gname
    return trees


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def build(xlsx_path, out_path):
    info(f"加载工作簿: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    matrix, col_groups = parse_prob_matrix(wb["Perk Group Probabilities"])
    backgrounds = parse_backgrounds(wb["Backgrounds"], matrix)
    traits = parse_traits(wb["Tree Multipliers"])
    attribute_stars = parse_attribute_stars(wb["Tree Multipliers"])
    exclusive_cascade = parse_exclusive_cascade(wb["Tree Multipliers"])
    self_weights = parse_self_weights(wb["Tree Multipliers"])
    projected = parse_projected_attributes(wb["Tree Multipliers"])
    perk_trees = parse_perks(wb["Perks"])

    # 组装 groups 定义
    groups_def = OrderedDict()
    for g in CANON_GROUPS_47:
        groups_def[g] = OrderedDict([
            ("category", GROUP_CATEGORY.get(g, "Unknown")),
            ("self_weight", self_weights.get(g)),
            ("order", list(GROUP_CATEGORY.keys()).index(g) if g in GROUP_CATEGORY else 999),
        ])
    # 补充 Tactician（矩阵无列但属 Shared）
    groups_def["Tactician"] = OrderedDict([
        ("category", "Shared"),
        ("self_weight", self_weights.get("Tactician")),
        ("order", list(GROUP_CATEGORY.keys()).index("Tactician")),
    ])

    data = OrderedDict([
        ("meta", OrderedDict([
            ("version", "0.7.6"),
            ("generated", "2026-06-29"),
            ("game", "Battle Brothers Reforged"),
        ])),
        ("config", OrderedDict([
            ("max_level", 11),
            ("skill_points", 10),
            ("melee_only_keyword", "Melee Only"),
            ("big_weight", BIG_WEIGHT),
            ("default_group_rolls", OrderedDict([
                ("Exclusive", 0.5),    # 0-1 范围，取期望 0.5
                ("Shared", 2),
                ("Weapon", 3),
                ("Armor", 2),
                ("Fighting Style", 2),
                ("Special", 0),
            ])),
        ])),
        ("groups", groups_def),
        ("backgrounds", backgrounds),
        ("traits", traits),
        ("exclusive_groups", exclusive_cascade),
        ("attribute_weights", attribute_stars),
        ("projected_attributes", projected),
        ("perk_trees", perk_trees),
        ("builds", OrderedDict()),  # 流派推荐待人工编写
    ])

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    info(f"已写出 data.json: {out_path}")
    info(f"  背景 {len(backgrounds)} / 特性 {len(traits)} / 组 {len(groups_def)} / 技能树 {len(perk_trees)}")

    # 数据质量报告
    print("\n========== 数据质量报告 ==========")
    print(f"告警总数: {len(WARNINGS)}")
    bg_ids = set(backgrounds.keys())
    matrix_ids = set(matrix.keys())
    only_matrix = matrix_ids - bg_ids
    only_bg = bg_ids - matrix_ids
    if only_matrix:
        print(f"[校验] 概率矩阵有但背景表没有的 ID: {sorted(only_matrix)}")
    if only_bg:
        print(f"[校验] 背景表有但概率矩阵没有的 ID(基础概率将为空): {sorted(only_bg)}")
    all_groups_referenced = set()
    for b in backgrounds.values():
        all_groups_referenced.update(b["multipliers"].keys())
        all_groups_referenced.update(b["base_probabilities"].keys())
    for t in traits.values():
        all_groups_referenced.update(t["weights"].keys())
    unknown = all_groups_referenced - set(groups_def.keys())
    if unknown:
        print(f"[校验] 被引用但未在 groups 定义的组: {sorted(unknown)}")
    else:
        print("[校验] 所有被引用组名均在 groups 中定义 ✓")
    tac_refs = sum(1 for b in backgrounds.values() if "Tactician" in b["multipliers"])
    print(f"[校验] Tactician 被背景 multipliers 引用: {tac_refs} 个背景(基础概率为 null,靠权重合成)")
    bad_probs = []
    for bg, probs in matrix.items():
        for g, p in probs.items():
            if p < 0 or p > 1.0001:
                bad_probs.append((bg, g, p))
    if bad_probs:
        print(f"[校验] 越界概率(>1 或 <0): {len(bad_probs)} 处, 示例: {bad_probs[:3]}")
    else:
        print("[校验] 概率值均在 [0,1] 范围内 ✓")

    log_path = os.path.splitext(out_path)[0] + "_parse_log.txt"
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(WARNINGS) if WARNINGS else "(无告警)")
    info(f"解析日志: {log_path}")
    return data


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    xlsx = sys.argv[1] if len(sys.argv) > 1 else os.path.join(here, "bb.xlsx")
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(here, "data.json")
    if not os.path.exists(xlsx):
        for cand in [
            os.path.join(here, "Battle Brothers Reforged - 0.7.6.xlsx"),
            "/sessions/serene-admiring-thompson/mnt/uploads/Battle Brothers Reforged - 0.7.6.xlsx",
        ]:
            if os.path.exists(cand):
                xlsx = cand
                break
    build(xlsx, out)
