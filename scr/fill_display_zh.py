# -*- coding: utf-8 -*-
"""
fill_display_zh.py
从 重铸手册0.6.8粗汉化.xlsx 提取中文名，填充到 data.json 的 display_zh 字段。

覆盖范围：
  - 背景：从「背景（重要）」sheet 提取 98 条映射（87 直接匹配 + 11 模糊匹配）
  - 特性：手工建立的 76 条中英对照（汉化册无英文 ID 列）
  - 技能树组：手工翻译的 49 条对照（汉化册标注「暂不翻译」）

用法：
  python fill_display_zh.py [--dry-run]
  --dry-run  仅打印变更，不实际写入 data.json
"""

import json
import os
import re
import sys
from collections import OrderedDict

import openpyxl

# ==============================
# 一、特性中英映射（76 条，从汉化册「特技树倍率」sheet 手工对照）
# ==============================
TRAIT_ZH_MAP = {
    "Addict": "药物成瘾",
    "Ailing": "虚弱",
    "Arena Fighter": "竞技场战士",
    "Arena Pit Fighter": "土坑斗士",
    "Arena Veteran": "竞技场老兵",
    "Asthmatic": "哮喘",
    "Athletic": "健壮",
    "Bleeder": "出血体质",
    "Bloodthirsty": "嗜血",
    "Brave": "勇敢",
    "Bright": "聪明",
    "Brute": "粗野",
    "Clubfooted": "畸形足",
    "Clumsy": "笨拙",
    "Cocky": "自大",
    "Craven": "懦夫",
    "Cultist Acolyte": "邪教侍僧",
    "Cultist Chosen": "邪教神选",
    "Cultist Disciple": "邪教门徒",
    "Cultist Fanatic": "邪教狂热",
    "Cultist Prophet": "邪教先知",
    "Cultist Zealot": "邪教狂信",
    "Dastard": "畏缩",
    "Deathwish": "死愿",
    "Determined": "坚定",
    "Dexterous": "灵巧",
    "Disloyal": "不忠",
    "Drunkard": "酒鬼",
    "Dumb": "蠢",
    "Eagle Eyes": "鹰眼",
    "Fainthearted": "懦弱",
    "Fat": "胖",
    "Fear Beasts": "怕野兽",
    "Fear Greenskins": "怕绿皮",
    "Fear Nobles": "怕贵族",
    "Fear Undead": "怕亡灵",
    "Fearless": "无畏",
    "Fragile": "脆弱",
    "Glorious Endurance": "坚忍之誉",
    "Glorious Quickness": "快速之誉",
    "Glorious Resolve": "决心之誉",
    "Gluttonous": "贪吃",
    "Greedy": "贪婪",
    "Hate Beasts": "恨野兽",
    "Hate Greenskins": "恨绿皮",
    "Hate Nobles": "恨贵族",
    "Hate Undead": "恨亡灵",
    "Hesitant": "犹豫",
    "Huge Trait": "巨大",
    "Impatient": "没耐心",
    "Insecure": "不安",
    "Intensive Training": "密集训练",
    "Iron Jaw": "铁颌",
    "Iron Lungs": "铁肺",
    "Irrational": "非理性",
    "Loyal": "忠诚",
    "Lucky": "幸运",
    "Mad": "疯狂",
    "Night Blind": "夜盲",
    "Night Owl": "猫头鹰",
    "Old": "老迈",
    "Optimist": "乐观",
    "Paranoid": "偏执",
    "Pessimist": "消极",
    "Player Character": "玩家角色",
    "Quick": "快速",
    "Short Sighted": "近视",
    "Spartan": "斯巴达",
    "Superstitious": "迷信",
    "Sure Footing": "下盘稳固",
    "Survivor": "幸存者",
    "Swift": "迅捷",
    "Teamplayer": "团队",
    "Tiny": "矮小",
    "Tough": "坚韧",
    "Weasel": "鼬鼠",
}

# ==============================
# 二、技能树组中英映射（48 组，自行翻译）
# ==============================
GROUP_ZH_MAP = {
    # Shared
    "Agile": "灵活",
    "Fast": "快速",
    "Tough": "坚韧",
    "Trained": "训练有素",
    "Unstoppable": "不可阻挡",
    "Vicious": "凶狠",
    "Vigorous": "精力充沛",
    "Tactician": "战术家",
    # Weapon
    "Axe": "斧",
    "Bow": "弓",
    "Cleaver": "砍刀",
    "Crossbow": "弩",
    "Dagger": "匕首",
    "Flail": "链枷",
    "Hammer": "锤",
    "Mace": "骨朵",
    "Polearm": "长柄",
    "Spear": "矛",
    "Sword": "剑",
    "Throwing": "投掷",
    # Armor
    "Heavy Armor": "重甲",
    "Light Armor": "轻甲",
    "Medium Armor": "中甲",
    # Fighting Style
    "Swift": "迅捷打击",
    "Power": "猛力打击",
    "Ranged": "远程战斗风格",
    "Shield": "盾牌战斗风格",
    # Exclusive
    "Knave": "奸猾",
    "Laborer": "劳工",
    "Militia": "民兵",
    "Noble": "贵族",
    "Pauper": "乞丐",
    "Raider": "掠袭者",
    "Soldier": "士兵",
    "Swordmaster": "剑术大师",
    "Trapper": "陷阱师",
    "Wildling": "野人",
    # Special
    "Back To Basics": "返朴归真",
    "Discovered Talent": "发掘潜能",
    "Fencer": "击剑手",
    "Gifted": "天才",
    "Leadership": "领袖",
    "Man Of Steel": "钢铁之躯",
    "Marksmanship": "神射手",
    "Professional": "专业人士",
    "Rising Star": "冉冉新星",
    "Student": "学生",
    # Always
    "General": "通用",
}

# ==============================
# 三、背景中文名 — 从汉化册加载 + 模糊匹配
# ==============================
HANHUA_BG_SHEET = "背景（重要）"

# 汉化册 ID 中的中文词 → 英文词替换规则（用于模糊匹配）
BG_FIX_RULES = [
    ("贵族", "noble"),
    ("士兵", "soldier"),
    ("民兵", "militia"),
    ("掠袭者", "raider"),
    ("远程战斗风格", "ranged"),
    ("sell剑", "sellsword"),
    ("弓yer", "bowyer"),
    ("剑master", "swordmaster"),
]

# 手工补充：汉化册没有但 data.json 有的背景（12 个中的例外）
BG_MANUAL_ZH = {
    "rf_old_swordmaster_background": "老剑圣",
    "rf_renowned_swordmaster_background": "传奇剑圣",
}


def fix_bg_id(hanhua_id: str) -> str:
    """将汉化册的 ID（可能含中文）修正为 data.json 背景 ID。"""
    fixed = hanhua_id
    for cn, en in BG_FIX_RULES:
        fixed = fixed.replace(cn, en)
    return fixed


def load_bg_zh_map(hanhua_path: str) -> dict:
    """从汉化册「背景（重要）」sheet 加载 背景ID→中文名 映射。"""
    wb = openpyxl.load_workbook(hanhua_path, data_only=True)
    ws = wb[HANHUA_BG_SHEET]
    raw_map = {}
    for r in range(3, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        if c1 and c2:
            raw_map[str(c1).strip()] = str(c2).strip()
    print(f"[背景] 从汉化册读取 {len(raw_map)} 条原始映射")
    return raw_map


def build_bg_zh_map(hanhua_path: str, data_bg_ids: list) -> dict:
    """构建 汉化册→data.json 的背景中文名映射，含模糊匹配。"""
    raw = load_bg_zh_map(hanhua_path)
    result = {}
    matched = 0
    fuzzy_matched = 0
    manual_filled = 0

    for bg_id in data_bg_ids:
        # 直接匹配
        if bg_id in raw:
            result[bg_id] = raw[bg_id]
            matched += 1
            continue
        # 模糊匹配：修正汉化册 key 后比对
        for hh_id, zh_name in raw.items():
            if fix_bg_id(hh_id) == bg_id:
                result[bg_id] = zh_name
                fuzzy_matched += 1
                print(f"  [模糊匹配] {hh_id} → {bg_id} → {zh_name}")
                break
        else:
            # 手工补充
            if bg_id in BG_MANUAL_ZH:
                result[bg_id] = BG_MANUAL_ZH[bg_id]
                manual_filled += 1
                print(f"  [手工补充] {bg_id} → {BG_MANUAL_ZH[bg_id]}")
            else:
                print(f"  [未找到] {bg_id} — 无中文名")

    print(f"[背景] 直接匹配: {matched}, 模糊匹配: {fuzzy_matched}, 手工补充: {manual_filled}, 未找到: {len(data_bg_ids) - matched - fuzzy_matched - manual_filled}")
    return result


# ==============================
# 四、主逻辑
# ==============================
def fill(data_json_path: str, hanhua_path: str, dry_run: bool = False):
    # 加载 data.json
    with open(data_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    stats = {"bg": 0, "trait": 0, "group": 0}

    # --- 背景 ---
    bg_ids = list(data["backgrounds"].keys())
    bg_zh = build_bg_zh_map(hanhua_path, bg_ids)
    for bg_id, zh_name in bg_zh.items():
        data["backgrounds"][bg_id]["display_zh"] = zh_name
        stats["bg"] += 1

    # --- 特性 ---
    for tid, tdata in data["traits"].items():
        if tid in TRAIT_ZH_MAP:
            tdata["display_zh"] = TRAIT_ZH_MAP[tid]
            stats["trait"] += 1
        else:
            print(f"  [特性未映射] {tid}")

    # --- 技能树组 ---
    for gid, gdata in data["groups"].items():
        if gid in GROUP_ZH_MAP:
            gdata["display_zh"] = GROUP_ZH_MAP[gid]
            stats["group"] += 1
        else:
            print(f"  [组未映射] {gid}")

    print(f"\n填充完成: 背景 {stats['bg']}/{len(bg_ids)}, 特性 {stats['trait']}/{len(data['traits'])}, 组 {stats['group']}/{len(data['groups'])}")

    if dry_run:
        print("(dry-run 模式，未写入文件)")
        return

    # 写回
    bak_path = data_json_path + ".bak"
    if os.path.exists(data_json_path):
        if os.path.exists(bak_path):
            os.remove(bak_path)
        os.rename(data_json_path, bak_path)
        print(f"已备份原文件: {bak_path}")

    with open(data_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已写入: {data_json_path}")


if __name__ == "__main__":
    dry_run = "--dry-run" in sys.argv
    data_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")
    hanhua = os.path.join(os.path.dirname(os.path.abspath(__file__)), "重铸手册0.6.8粗汉化.xlsx")

    if not os.path.exists(data_json):
        print(f"错误: 未找到 {data_json}")
        sys.exit(1)
    if not os.path.exists(hanhua):
        print(f"错误: 未找到 {hanhua}")
        sys.exit(1)

    fill(data_json, hanhua, dry_run)
